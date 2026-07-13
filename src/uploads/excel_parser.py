from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from src.templates.excel_schema import SHEET_NAMES
from src.uploads.models import UploadParseError


def _rows_as_dicts(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        return []
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def parse_excel_upload(content: bytes) -> dict[str, list[dict[str, Any]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=False)
    except Exception as exc:
        raise UploadParseError("The Excel workbook could not be opened safely.") from exc

    missing = [name for name in SHEET_NAMES if name not in workbook.sheetnames]
    if missing:
        raise UploadParseError(f"Missing required sheets: {', '.join(missing)}")

    if getattr(workbook, "vba_archive", None) is not None:
        raise UploadParseError("Macro-enabled workbooks are not accepted.")

    return {name: _rows_as_dicts(workbook[name]) for name in SHEET_NAMES}
