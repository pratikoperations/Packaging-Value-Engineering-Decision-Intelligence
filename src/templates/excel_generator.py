from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from src.category_registry import default_registry

from .dropdown_values import DROPDOWNS
from .excel_schema import (
    COMMERCIAL_FIELDS,
    DOCUMENT_HEADERS,
    LOGISTICS_FIELDS,
    PROJECT_HEADERS,
    QUALITY_HEADERS,
    SHEET_NAMES,
)
from .instructions_builder import build_instructions

_HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
_REQUIRED_FILL = PatternFill("solid", fgColor="F4CCCC")
_RECOMMENDED_FILL = PatternFill("solid", fgColor="FCE5CD")
_OPTIONAL_FILL = PatternFill("solid", fgColor="E7E6E6")
_CATEGORY_ALIASES = {"corrugated_shipping_case": "corrugated"}


def _style_headers(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _apply_requirement_fill(ws, row: int, requirement: str) -> None:
    fill = {
        "mandatory": _REQUIRED_FILL,
        "recommended": _RECOMMENDED_FILL,
        "optional": _OPTIONAL_FILL,
    }.get(requirement)
    if fill:
        for cell in ws[row]:
            cell.fill = fill


def _add_list_validation(ws, column: int, values: tuple[str, ...], start: int = 2, end: int = 500) -> None:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"{ws.cell(1, column).column_letter}{start}:{ws.cell(1, column).column_letter}{end}")


def _write_value_rows(ws, rows) -> None:
    ws.append(PROJECT_HEADERS)
    for field in rows:
        unit = ", ".join(field.units)
        example = ""
        ws.append((field.key, field.label, field.requirement, "", unit, field.description, example, "", "", "", "", "not_reviewed"))
        _apply_requirement_fill(ws, ws.max_row, field.requirement)
    _style_headers(ws)
    _add_list_validation(ws, 3, DROPDOWNS["requirement"])
    _add_list_validation(ws, 8, DROPDOWNS["source_classification"])
    _add_list_validation(ws, 12, DROPDOWNS["validation_status"])


def _write_common_rows(ws, rows) -> None:
    ws.append(PROJECT_HEADERS)
    for key, label, requirement, unit, example in rows:
        ws.append((key, label, requirement, "", unit, "", example, "", "", "", "", "not_reviewed"))
        _apply_requirement_fill(ws, ws.max_row, requirement)
    _style_headers(ws)
    _add_list_validation(ws, 3, DROPDOWNS["requirement"])
    _add_list_validation(ws, 8, DROPDOWNS["source_classification"])
    _add_list_validation(ws, 12, DROPDOWNS["validation_status"])


def generate_workbook(category: str, objective: str, change_type: str) -> bytes:
    registry = default_registry()
    registry_category = _CATEGORY_ALIASES.get(category, category)
    definition = registry.get(registry_category)
    if not definition.supports_objective(objective):
        raise ValueError("Unsupported project objective.")
    if not definition.supports_change_type(change_type):
        raise ValueError("Unsupported change type for category.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_NAMES:
        workbook.create_sheet(name)

    instructions = workbook["INSTRUCTIONS"]
    instructions.append(("Topic", "Guidance"))
    for row in build_instructions(definition.display_name, objective, change_type):
        instructions.append(row)
    _style_headers(instructions)

    project = workbook["PROJECT"]
    project.append(PROJECT_HEADERS)
    project_rows = (
        ("project_code", "Project code", "mandatory", "text", "PVE-PACK-001"),
        ("project_name", "Project name", "mandatory", "text", "Packaging improvement"),
        ("category", "Packaging category", "mandatory", "text", category),
        ("objective", "Project objective", "mandatory", "text", objective),
        ("change_type", "Change type", "mandatory", "text", change_type),
        ("product_sku", "Product or SKU", "mandatory", "text", "SKU-001"),
        ("business_unit_plant", "Business unit or plant", "mandatory", "text", "Plant A"),
        ("project_owner", "Project owner", "mandatory", "text", "Owner"),
        ("currency", "Currency", "mandatory", "text", "INR"),
        ("volume_unit", "Volume unit", "mandatory", "text", "units/year"),
    )
    for key, label, requirement, unit, example in project_rows:
        project.append((key, label, requirement, "", unit, "", example, "", "", "", "", "not_reviewed"))
        _apply_requirement_fill(project, project.max_row, requirement)
    _style_headers(project)
    _add_list_validation(project, 8, DROPDOWNS["source_classification"])
    _add_list_validation(project, 12, DROPDOWNS["validation_status"])

    _write_value_rows(workbook["BASELINE"], definition.fields)
    _write_value_rows(workbook["PROPOSED"], definition.fields)
    _write_common_rows(workbook["COMMERCIAL"], COMMERCIAL_FIELDS)
    _write_common_rows(workbook["LOGISTICS"], LOGISTICS_FIELDS)

    quality = workbook["QUALITY_TESTS"]
    quality.append(QUALITY_HEADERS)
    for test in definition.tests:
        requirement = "mandatory" if test.critical else "recommended"
        quality.append((test.name, requirement, "yes" if test.critical else "no", "proposed", "", "", "", "", "", "", "not_reviewed", ""))
        _apply_requirement_fill(quality, quality.max_row, requirement)
    _style_headers(quality)
    _add_list_validation(quality, 4, DROPDOWNS["context"])
    _add_list_validation(quality, 7, DROPDOWNS["source_classification"])
    _add_list_validation(quality, 11, DROPDOWNS["validation_status"])

    documents = workbook["DOCUMENT_REGISTER"]
    documents.append(DOCUMENT_HEADERS)
    for index, document in enumerate(definition.documents, start=1):
        documents.append((f"DOC-{index:03d}", document.document_type, "", category, "baseline", "", "", "", "", "not_reviewed", document.requirement, "missing", ""))
        _apply_requirement_fill(documents, documents.max_row, document.requirement)
    _style_headers(documents)
    _add_list_validation(documents, 5, DROPDOWNS["context"])
    _add_list_validation(documents, 10, DROPDOWNS["verification_status"])
    _add_list_validation(documents, 11, DROPDOWNS["requirement"])
    _add_list_validation(documents, 12, DROPDOWNS["upload_status"])

    for ws in workbook.worksheets:
        for column in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
            ws.column_dimensions[column[0].column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
