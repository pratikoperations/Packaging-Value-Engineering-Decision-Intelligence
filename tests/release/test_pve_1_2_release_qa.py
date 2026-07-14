from __future__ import annotations

import json
import sqlite3
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.category_registry import (
    analyze_corrugated_economics,
    assess_evidence_confidence,
    build_engineering_recommendation,
    compare_simple_pallet_patterns,
    material_comparison,
    physical_sustainability_indicators,
    screen_corrugated,
    validate_tolerance,
)
from src.persistence import (
    Database,
    DatasetRepository,
    ProjectRepository,
    ReadinessRepository,
    TechnicalAssessmentRepository,
)
from src.persistence import migrations
from src.persistence.migrations import current_schema_version, initialize_database
from src.templates import generate_workbook
from src.uploads.normalizer import normalize_user_dataset


ROOT = Path(__file__).resolve().parents[2]
CASES_PATH = ROOT / "data" / "pve_1_2_corrugated_demonstration_cases.json"
PROHIBITED_AUTOMATIC_OUTCOMES = {"Approved", "Rejected", "Conditional"}


class PVE12ReleaseQATestCase(unittest.TestCase):
    def load_cases(self):
        return json.loads(CASES_PATH.read_text(encoding="utf-8"))

    def test_all_eight_cases_are_governed_synthetic_data(self):
        payload = self.load_cases()
        self.assertEqual(payload["dataset_type"], "synthetic_demonstration_data")
        self.assertIn("SYNTHETIC DEMONSTRATION DATA ONLY", payload["notice"])
        self.assertEqual(len(payload["cases"]), 8)
        self.assertEqual(len({case["case_id"] for case in payload["cases"]}), 8)
        self.assertTrue(all(case["synthetic"] is True for case in payload["cases"]))

    def test_demonstration_case_recommendation_outcomes_and_non_approval(self):
        for case in self.load_cases()["cases"]:
            scenario = case["scenario"]
            economics = {
                key: value
                for key, value in scenario.items()
                if key in {
                    "gross_annual_benefit", "incremental_failure_cost",
                    "risk_adjusted_annual_benefit", "incremental_working_capital",
                    "obsolete_stock_write_off", "first_year_net_benefit",
                }
            }
            material_logistics = {
                key: value
                for key, value in scenario.items()
                if key in {"annual_material_change_kg", "annual_pallet_movements_change"}
            }
            recommendation = build_engineering_recommendation(
                screening_outcome=scenario["screening_outcome"],
                technical_blockers=scenario.get("technical_blockers", ()),
                required_trials=scenario.get("required_trials", ()),
                evidence_confidence=scenario["evidence_confidence"],
                material_logistics=material_logistics,
                economics=economics,
            )
            self.assertEqual(recommendation.outcome, case["expected_outcome"], case["case_id"])
            self.assertNotIn(recommendation.outcome, PROHIBITED_AUTOMATIC_OUTCOMES)
            self.assertIn("not an approval decision", recommendation.limitations[0])

    def test_right_sized_case_end_to_end_through_immutable_assessment(self):
        project_input = {
            "project_id": "P-DEMO-01", "project_name": "Synthetic right-size",
            "category": "corrugated", "annual_volume": 100000, "currency": "INR",
        }
        tolerance = {
            "field_key": "internal_length_mm", "context": "proposed",
            "nominal": "400", "minimum": "398", "maximum": "402", "unit": "mm",
            "inspection_method": "calibrated steel rule", "criticality": "major",
            "source_classification": "manually_entered_fact",
            "source_reference": "SYNTHETIC-SPEC-2", "version": "2",
            "validation_status": "valid",
        }
        normalized = normalize_user_dataset(
            {
                "intake_values": [{
                    "field_key": "internal_length_mm", "context": "proposed",
                    "value": "400", "unit": "mm",
                    "source_classification": "uploaded_fact",
                    "evidence_reference": "SYNTHETIC-DRAWING-2",
                }],
                "specification_tolerances": [tolerance],
            },
            project_input,
        )
        self.assertEqual(normalized["intake_values"][0]["value"], 400)
        self.assertEqual(validate_tolerance(normalized["specification_tolerances"][0]), ())

        expected = {
            "project_id": "P-DEMO-01", "context": "proposed",
            "specification_version": "SPEC-2", "supplier_name": "Synthetic Supplier",
            "manufacturing_site": "Synthetic Site", "material_structure": "5-ply BC",
            "laboratory_name": "Synthetic Lab", "sample_or_batch_reference": "SYN-BATCH-1",
        }
        evidence = [
            {
                **expected, "evidence_id": "SYN-BCT-1", "test_method": "BCT",
                "result_value": 1400, "unit": "N", "source_classification": "laboratory_tested",
                "validation_status": "valid", "test_date": "2026-07-01", "valid_until": "2099-12-31",
            },
            {
                **expected, "evidence_id": "SYN-ECT-1", "test_method": "ECT",
                "result_value": 8.2, "unit": "kN/m", "source_classification": "laboratory_tested",
                "validation_status": "valid", "test_date": "2026-07-01", "valid_until": "2099-12-31",
            },
        ]
        confidence = assess_evidence_confidence(evidence)
        self.assertEqual(confidence.classification, "High evidence confidence")

        screening = screen_corrugated(
            requirements={
                "compression_requirement_n": 1200, "ect_requirement_kn_m": 8,
                "stack_layers_required": 5, "proposed_stack_layers": 5,
                "pallet_load_kg": 650, "maximum_pallet_weight_kg": 800,
                "stacking_mode": "static", "storage_duration_days": 30,
                "external_length_mm": 400, "external_width_mm": 300,
                "external_height_mm": 250,
            },
            evidence=evidence,
            expected_evidence_context=expected,
        )
        self.assertEqual(screening.outcome, "criteria met")
        self.assertFalse(screening.blockers)

        material = material_comparison(
            annual_volume_cases=100000,
            baseline={"case_weight_g": 500, "blank_length_mm": 1000, "blank_width_mm": 600},
            proposed={"case_weight_g": 450, "blank_length_mm": 950, "blank_width_mm": 580},
        )
        pallets = compare_simple_pallet_patterns({
            "case_external_length_mm": 400, "case_external_width_mm": 300,
            "case_external_height_mm": 250, "case_weight_kg": 10,
            "pallet_length_mm": 1200, "pallet_width_mm": 1000,
            "pallet_height_limit_mm": 1500, "pallet_weight_limit_kg": 1000,
            "empty_pallet_weight_kg": 25, "validated_stack_layers": 5,
            "annual_volume_cases": 100000,
        })
        sustainability = physical_sustainability_indicators(
            annual_volume_cases=100000,
            baseline={"case_weight_g": 500},
            proposed={
                "case_weight_g": 450, "product_weight_per_case_kg": 12,
                "recycled_content_percent": 80, "virgin_fibre_percent": 20,
            },
            pallet_movements_baseline=10000,
            pallet_movements_proposed=min(item.annual_pallet_movements for item in pallets),
        )
        self.assertEqual(material["annual_material_change"].value, -5000)
        self.assertGreater(max(item.cases_per_pallet for item in pallets), 0)
        self.assertEqual(sustainability["annual_paper_reduction"].value, 5000)

        should_cost = []
        for context, components in {
            "baseline": {"board_or_paper": 12, "conversion": 3, "printing": 1, "freight": 2},
            "proposed": {"board_or_paper": 10, "conversion": 3, "printing": 1, "freight": 1.5},
        }.items():
            for component, value in components.items():
                should_cost.append({
                    "record_id": f"{context}-{component}", "context": context,
                    "component": component, "value_per_case": value, "currency": "INR",
                    "source_classification": "manually_entered_fact",
                    "source_reference": "SYNTHETIC-COST-1",
                })
        failure = [
            {"record_id": "fail-b", "context": "baseline", "annual_cases": 100000,
             "damage_rate_percent": 1, "loss_per_damaged_case": 50, "currency": "INR",
             "source_classification": "uploaded_fact", "source_reference": "SYNTHETIC-HISTORY"},
            {"record_id": "fail-p", "context": "proposed", "annual_cases": 100000,
             "damage_rate_percent": 1, "loss_per_damaged_case": 50, "currency": "INR",
             "source_classification": "assumption", "source_reference": "SYNTHETIC-ASSUMPTION"},
        ]
        inventory = [
            {"record_id": "inv-b", "context": "baseline", "annual_cases": 100000,
             "inventory_days": 20, "unit_inventory_value": 18, "transition_stock_units": 0,
             "obsolete_stock_units": 0, "write_off_percent": 100, "currency": "INR",
             "source_classification": "uploaded_fact", "source_reference": "SYNTHETIC-ERP"},
            {"record_id": "inv-p", "context": "proposed", "annual_cases": 100000,
             "inventory_days": 20, "unit_inventory_value": 15.5, "transition_stock_units": 0,
             "obsolete_stock_units": 0, "write_off_percent": 100, "currency": "INR",
             "source_classification": "manually_entered_fact", "source_reference": "SYNTHETIC-PLAN"},
        ]
        economics = analyze_corrugated_economics(
            currency="INR", annual_cases=100000,
            should_cost_inputs=should_cost, failure_cost_inputs=failure,
            inventory_inputs=inventory,
            one_time_costs=[{
                "record_id": "trial", "component": "trials", "value": 25000,
                "currency": "INR", "source_classification": "manually_entered_fact",
                "source_reference": "SYNTHETIC-TRIAL",
            }],
        )
        self.assertGreater(economics.outputs["risk_adjusted_annual_benefit"].value, 0)

        recommendation = build_engineering_recommendation(
            screening_outcome=screening.outcome,
            technical_blockers=screening.blockers,
            evidence_confidence=confidence.classification,
            material_logistics={
                "annual_material_change_kg": material["annual_material_change"].value,
                "best_cases_per_pallet": max(item.cases_per_pallet for item in pallets),
            },
            economics={
                "risk_adjusted_annual_benefit": economics.outputs["risk_adjusted_annual_benefit"].value,
            },
        )
        self.assertEqual(recommendation.outcome, "criteria met for engineering review")

        with tempfile.TemporaryDirectory() as tempdir:
            database = Database(Path(tempdir) / "pve.sqlite3")
            initialize_database(database)
            projects = ProjectRepository(database)
            datasets = DatasetRepository(database)
            readiness = ReadinessRepository(database)
            assessments = TechnicalAssessmentRepository(database)
            project = projects.create(
                project_id="P-DEMO-01", project_code="PVE12-DEMO-01",
                project_name="Synthetic right-size", category="corrugated",
                currency="INR", annual_volume=100000,
            )
            dataset = datasets.create_version(
                project_id=project["project_id"], source_type="json",
                canonical_data=normalized, validation_status="valid",
            )
            readiness_record = readiness.create(
                project_id=project["project_id"], dataset_id=dataset["dataset_id"],
                assessment={"score_percent": 95, "stage": "validation_ready"},
            )
            saved = assessments.create(
                project_id=project["project_id"],
                readiness_assessment_id=readiness_record["readiness_assessment_id"],
                dataset_id=dataset["dataset_id"],
                baseline_specification_version="BASE-1",
                proposed_specification_version="SPEC-2",
                rule_set_version="PVE-1.2-B8",
                threshold_profile_id=None,
                threshold_references=["SYNTHETIC-BCT-REQ"],
                evidence_references=[{"evidence_id": row["evidence_id"], "project_id": project["project_id"]} for row in evidence],
                formula_inputs={"compression_requirement_n": 1200},
                assumptions=["Synthetic demonstration data only."],
                technical_outcomes=dict(recommendation.technical_outcomes),
                commercial_outcomes=dict(recommendation.commercial_outcomes),
                blockers=list(recommendation.blockers),
                required_trials=list(recommendation.required_trials),
                evidence_confidence_status=recommendation.evidence_confidence,
                recommendation_outcome=recommendation.outcome,
            )
            self.assertEqual(saved["recommendation_outcome"], "criteria met for engineering review")
            with self.assertRaisesRegex(ValueError, "immutable"):
                assessments.update(saved["technical_assessment_id"])
            with self.assertRaisesRegex(ValueError, "immutable"):
                assessments.delete(saved["technical_assessment_id"])

    def test_additive_migration_from_schema_versions_one_two_and_three(self):
        for starting_version in (1, 2, 3):
            with self.subTest(starting_version=starting_version), tempfile.TemporaryDirectory() as tempdir:
                database = Database(Path(tempdir) / "pve.sqlite3")
                with database.transaction() as connection:
                    connection.executescript(migrations._BASE_SCHEMA)
                    connection.execute("INSERT OR IGNORE INTO schema_migrations(version) VALUES (1)")
                    if starting_version >= 2:
                        migrations._apply_v2(connection)
                    if starting_version >= 3:
                        migrations._apply_v3(connection)
                self.assertEqual(current_schema_version(database), starting_version)
                self.assertEqual(initialize_database(database), 4)
                self.assertEqual(current_schema_version(database), 4)
                with database.connect() as connection:
                    tables = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")}
                self.assertIn("technical_assessments", tables)

    def test_all_immutable_record_families_have_update_and_delete_triggers(self):
        with tempfile.TemporaryDirectory() as tempdir:
            database = Database(Path(tempdir) / "pve.sqlite3")
            initialize_database(database)
            with database.connect() as connection:
                triggers = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='trigger'")}
            families = (
                "project_datasets", "threshold_profiles", "scenarios",
                "decision_snapshots", "readiness_assessments", "technical_assessments",
            )
            for family in families:
                self.assertIn(f"{family}_immutable_update", triggers)
                self.assertIn(f"{family}_immutable_delete", triggers)

    def test_archived_project_and_cross_project_assessment_references_are_rejected(self):
        with tempfile.TemporaryDirectory() as tempdir:
            database = Database(Path(tempdir) / "pve.sqlite3")
            initialize_database(database)
            projects = ProjectRepository(database)
            datasets = DatasetRepository(database)
            assessments = TechnicalAssessmentRepository(database)
            first = projects.create(project_code="P1", project_name="First", category="corrugated", currency="INR", annual_volume=100)
            second = projects.create(project_code="P2", project_name="Second", category="corrugated", currency="INR", annual_volume=100)
            second_dataset = datasets.create_version(project_id=second["project_id"], source_type="json", canonical_data={"v": 1}, validation_status="valid")
            common = dict(
                project_id=first["project_id"], readiness_assessment_id=None,
                dataset_id=second_dataset["dataset_id"], baseline_specification_version="B1",
                proposed_specification_version="P1", rule_set_version="PVE-1.2-B8",
                threshold_profile_id=None, threshold_references=[], evidence_references=[],
                formula_inputs={}, assumptions=[], technical_outcomes={}, commercial_outcomes={},
                blockers=[], required_trials=[], evidence_confidence_status="Not assessable",
                recommendation_outcome="insufficient technical data",
            )
            with self.assertRaisesRegex(ValueError, "same project"):
                assessments.create(**common)
            projects.archive(first["project_id"])
            with self.assertRaisesRegex(ValueError, "read-only"):
                assessments.create(**dict(common, dataset_id=None))

    def test_json_csv_compatible_normalization_and_excel_template_regression(self):
        project = {"project_id": "P-QA", "project_name": "QA", "category": "corrugated", "annual_volume": 1000, "currency": "INR"}
        normalized = normalize_user_dataset(
            {
                "pallet_pattern_inputs": [{"pallet_length_mm": "1200", "annual_volume_cases": "10000"}],
                "failure_cost_inputs": [{"damage_rate_percent": "1.2", "loss_per_damaged_case": "50"}],
                "specification_tolerances": [{
                    "field_key": "internal_length_mm", "context": "proposed",
                    "nominal": "400", "minimum": "398", "maximum": "402",
                    "unit": "mm", "inspection_method": "rule", "criticality": "major",
                    "source_classification": "manually_entered_fact", "source_reference": "SYN-QA",
                    "version": "1", "validation_status": "valid",
                }],
            },
            project,
        )
        self.assertEqual(normalized["pallet_pattern_inputs"][0]["pallet_length_mm"], 1200)
        self.assertEqual(normalized["failure_cost_inputs"][0]["damage_rate_percent"], 1.2)
        workbook = load_workbook(BytesIO(generate_workbook("corrugated", "Cost reduction", "Size optimization")))
        self.assertTrue({"BASELINE", "PROPOSED"}.issubset(workbook.sheetnames))
        fields = {workbook["PROPOSED"].cell(row, 1).value for row in range(2, workbook["PROPOSED"].max_row + 1)}
        self.assertIn("box_style", fields)
        self.assertIn("blank_length_mm", fields)


if __name__ == "__main__":
    unittest.main()
