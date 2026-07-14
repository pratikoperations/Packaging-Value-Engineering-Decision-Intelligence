from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from src.category_registry import default_registry
from src.templates import SHEET_NAMES, generate_workbook


class ExcelGeneratorTestCase(unittest.TestCase):
    def test_all_eight_categories_generate_expected_workbook(self):
        registry = default_registry()
        for category in registry.keys():
            with self.subTest(category=category):
                definition = registry.get(category)
                content = generate_workbook(category, definition.objectives[0], definition.change_types[0])
                workbook = load_workbook(BytesIO(content))
                self.assertEqual(tuple(workbook.sheetnames), SHEET_NAMES)
                self.assertFalse(workbook.vba_archive)
                self.assertGreater(workbook["BASELINE"].max_row, 1)
                self.assertGreater(workbook["PROPOSED"].max_row, 1)
                self.assertGreater(workbook["QUALITY_TESTS"].max_row, 1)
                self.assertGreater(workbook["DOCUMENT_REGISTER"].max_row, 1)

    def test_objective_and_change_type_are_written_to_instructions(self):
        content = generate_workbook("labels", "Cost reduction", "Substrate change")
        workbook = load_workbook(BytesIO(content), data_only=True)
        values = {row[0]: row[1] for row in workbook["INSTRUCTIONS"].iter_rows(min_row=2, values_only=True)}
        self.assertEqual(values["Project objective"], "Cost reduction")
        self.assertEqual(values["Change type"], "Substrate change")

    def test_invalid_change_type_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "Unsupported change type"):
            generate_workbook("glass", "Cost reduction", "Ply reduction")

    def test_required_metadata_columns_exist(self):
        content = generate_workbook("corrugated", "Cost reduction", "GSM reduction")
        workbook = load_workbook(BytesIO(content))
        baseline_headers = [cell.value for cell in workbook["BASELINE"][1]]
        for header in (
            "requirement", "unit", "source_classification", "evidence_reference",
            "supplier_name", "test_date", "validation_status",
        ):
            self.assertIn(header, baseline_headers)

    def test_dropdown_validation_and_requirement_colours_exist(self):
        content = generate_workbook("rigid_plastic", "Material reduction", "Weight reduction")
        workbook = load_workbook(BytesIO(content))
        baseline = workbook["BASELINE"]
        self.assertGreater(len(baseline.data_validations.dataValidation), 0)
        self.assertNotEqual(baseline["A2"].fill.fgColor.rgb, "00000000")


if __name__ == "__main__":
    unittest.main()
