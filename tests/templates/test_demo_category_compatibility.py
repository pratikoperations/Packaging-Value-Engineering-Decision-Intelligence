from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from src.templates import generate_workbook


class DemoCategoryWorkbookCompatibilityTests(unittest.TestCase):
    def test_legacy_demo_category_generates_workbook(self) -> None:
        content = generate_workbook(
            "corrugated_shipping_case",
            "Cost reduction",
            "Size optimization",
        )

        workbook = load_workbook(BytesIO(content), read_only=True)
        self.assertIn("PROJECT", workbook.sheetnames)
        self.assertIn("INSTRUCTIONS", workbook.sheetnames)

    def test_canonical_corrugated_category_still_generates_workbook(self) -> None:
        content = generate_workbook(
            "corrugated",
            "Cost reduction",
            "Size optimization",
        )
        self.assertGreater(len(content), 0)

    def test_unknown_category_still_fails(self) -> None:
        with self.assertRaisesRegex(KeyError, "Unknown packaging category"):
            generate_workbook("unknown", "Cost reduction", "Size optimization")

    def test_invalid_objective_still_fails(self) -> None:
        with self.assertRaisesRegex(ValueError, "Unsupported project objective"):
            generate_workbook(
                "corrugated_shipping_case",
                "Unsupported objective",
                "Size optimization",
            )

    def test_invalid_change_type_still_fails(self) -> None:
        with self.assertRaisesRegex(ValueError, "Unsupported change type"):
            generate_workbook(
                "corrugated_shipping_case",
                "Cost reduction",
                "Unsupported change",
            )


if __name__ == "__main__":
    unittest.main()
