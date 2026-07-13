from __future__ import annotations

import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.application import ProjectService
from src.persistence import Database, DatasetRepository, ProjectRepository
from src.persistence.migrations import initialize_database
from src.templates.excel_generator import generate_workbook
from src.uploads.models import UploadParseError
from src.uploads.service import UploadService


class ExcelUploadTestCase(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        database = Database(Path(self.tempdir.name) / "pve.sqlite3")
        initialize_database(database)
        projects = ProjectService(ProjectRepository(database))
        self.project = projects.create_project(
            project_code="PVE-XLSX-1",
            project_name="Excel intake",
            category="corrugated",
            objective="Cost reduction",
            change_type="GSM reduction",
            currency="INR",
            annual_volume=1000,
            volume_unit="units/year",
        )
        self.service = UploadService(DatasetRepository(database))

    def tearDown(self):
        self.tempdir.cleanup()

    def _completed_workbook(self) -> bytes:
        data = generate_workbook("corrugated", "Cost reduction", "GSM reduction")
        workbook = load_workbook(BytesIO(data))
        project_values = {
            "project_code": "PVE-XLSX-1", "project_name": "Excel intake",
            "category": "corrugated", "objective": "Cost reduction",
            "change_type": "GSM reduction", "product_sku": "SKU-1",
            "business_unit_plant": "Plant A", "project_owner": "Owner",
            "currency": "INR", "volume_unit": "units/year",
        }
        for row in workbook["PROJECT"].iter_rows(min_row=2):
            key = row[0].value
            row[3].value = project_values.get(key)
            row[7].value = "manually_entered_fact"
        for sheet in ("BASELINE", "PROPOSED"):
            for row in workbook[sheet].iter_rows(min_row=2):
                requirement = row[2].value
                row[3].value = 1 if requirement == "mandatory" else None
                row[7].value = "manually_entered_fact" if row[3].value is not None else None
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_valid_workbook_is_prepared_and_saved(self):
        prepared = self.service.prepare_excel(
            content=self._completed_workbook(), filename="corrugated.xlsx", project=self.project
        )
        self.assertTrue(prepared.validation.is_valid, prepared.validation.issues)
        saved = self.service.save_valid_dataset(project_id=self.project["project_id"], prepared=prepared)
        self.assertEqual(saved["source_type"], "excel_template")

    def test_missing_sheet_is_rejected(self):
        workbook = load_workbook(BytesIO(generate_workbook("corrugated", "Cost reduction", "GSM reduction")))
        del workbook["PROPOSED"]
        output = BytesIO(); workbook.save(output)
        with self.assertRaises(UploadParseError):
            self.service.prepare_excel(content=output.getvalue(), filename="bad.xlsx", project=self.project)

    def test_category_mismatch_is_invalid(self):
        workbook = load_workbook(BytesIO(self._completed_workbook()))
        for row in workbook["PROJECT"].iter_rows(min_row=2):
            if row[0].value == "category":
                row[3].value = "glass"
        output = BytesIO(); workbook.save(output)
        prepared = self.service.prepare_excel(content=output.getvalue(), filename="bad.xlsx", project=self.project)
        self.assertFalse(prepared.validation.is_valid)
        self.assertIn("category_mismatch", {issue.code for issue in prepared.validation.issues})

    def test_invalid_source_classification_blocks_save(self):
        workbook = load_workbook(BytesIO(self._completed_workbook()))
        workbook["BASELINE"]["H2"] = "verified"
        output = BytesIO(); workbook.save(output)
        prepared = self.service.prepare_excel(content=output.getvalue(), filename="bad.xlsx", project=self.project)
        self.assertFalse(prepared.validation.is_valid)
        with self.assertRaisesRegex(ValueError, "Invalid uploads"):
            self.service.save_valid_dataset(project_id=self.project["project_id"], prepared=prepared)

    def test_invalid_numeric_value_and_unit_are_reported(self):
        workbook = load_workbook(BytesIO(self._completed_workbook()))
        workbook["BASELINE"]["D2"] = "not-a-number"
        workbook["BASELINE"]["E2"] = "inch"
        output = BytesIO(); workbook.save(output)
        prepared = self.service.prepare_excel(content=output.getvalue(), filename="bad.xlsx", project=self.project)
        codes = {issue.code for issue in prepared.validation.issues}
        self.assertIn("invalid_number", codes)
        self.assertIn("invalid_unit", codes)


if __name__ == "__main__":
    unittest.main()
