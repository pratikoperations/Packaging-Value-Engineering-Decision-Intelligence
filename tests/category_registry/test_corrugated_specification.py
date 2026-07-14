import unittest

from src.category_registry import (
    BOX_STYLES,
    CONVERTING_PROFILES,
    compare_specifications,
    default_registry,
    validate_tolerance,
    validate_tolerances,
)
from src.templates import generate_workbook
from src.uploads.normalizer import normalize_user_dataset


class CorrugatedSpecificationTestCase(unittest.TestCase):
    def _valid_tolerance(self):
        return {
            "field_key": "internal_length_mm",
            "context": "proposed",
            "nominal": 400,
            "minimum": 398,
            "maximum": 402,
            "unit": "mm",
            "inspection_method": "calibrated steel rule",
            "criticality": "major",
            "source_classification": "manually_entered_fact",
            "source_reference": "SPEC-REV-B",
            "version": "2",
            "validation_status": "valid",
        }

    def test_corrugated_registry_contains_build_2_fields(self):
        keys = {field.key for field in default_registry().get("corrugated").fields}
        expected = {
            "box_style", "converting_profile",
            "internal_length_mm", "internal_width_mm", "internal_height_mm",
            "external_length_mm", "external_width_mm", "external_height_mm",
            "flute_combination", "paper_layer_structure", "layer_gsm_profile",
            "board_caliper_mm", "joint_type", "closure_method", "print_process",
            "print_colour_count", "coating_or_treatment", "blank_length_mm",
            "blank_width_mm", "gross_packed_weight_kg", "case_pack_quantity",
            "artwork_revision", "regulatory_markings",
        }
        self.assertTrue(expected.issubset(keys))

    def test_box_style_and_converting_profiles_are_configuration_driven(self):
        self.assertIn("regular_slotted_container", BOX_STYLES)
        self.assertIn("die_cut_case", BOX_STYLES)
        self.assertIn("slotted_glued", CONVERTING_PROFILES)
        self.assertIn("die_cut_stitched", CONVERTING_PROFILES)

    def test_valid_sourced_tolerance_passes(self):
        self.assertEqual(validate_tolerance(self._valid_tolerance()), ())

    def test_tolerance_minimum_cannot_exceed_maximum(self):
        record = self._valid_tolerance()
        record["minimum"], record["maximum"] = 405, 402
        self.assertIn("Tolerance minimum cannot exceed maximum", validate_tolerance(record))

    def test_nominal_must_be_within_tolerance(self):
        record = self._valid_tolerance()
        record["nominal"] = 410
        self.assertIn("Tolerance nominal must be within minimum and maximum", validate_tolerance(record))

    def test_unsourced_tolerance_is_rejected(self):
        record = self._valid_tolerance()
        record["source_reference"] = ""
        record["version"] = ""
        issues = validate_tolerance(record)
        self.assertIn("Tolerance source_reference is required", issues)
        self.assertIn("Tolerance version is required", issues)

    def test_invalid_source_classification_is_rejected(self):
        record = self._valid_tolerance()
        record["source_classification"] = "verified"
        self.assertIn("Tolerance source_classification is invalid", validate_tolerance(record))

    def test_duplicate_tolerance_versions_are_rejected(self):
        record = self._valid_tolerance()
        issues = validate_tolerances((record, dict(record)))
        self.assertTrue(any("Duplicate tolerance record" in issue for issue in issues))

    def test_baseline_proposed_comparison_is_transparent(self):
        baseline = ({"field_key": "box_style", "value": "regular_slotted_container", "unit": ""},)
        proposed = ({"field_key": "box_style", "value": "die_cut_case", "unit": ""},)
        differences = compare_specifications(baseline, proposed)
        self.assertEqual(len(differences), 1)
        self.assertEqual(differences[0].field_key, "box_style")
        self.assertEqual(differences[0].baseline_value, "regular_slotted_container")
        self.assertEqual(differences[0].proposed_value, "die_cut_case")

    def test_normalizer_retains_traceability_and_coerces_tolerances(self):
        raw = {
            "intake_values": [{
                "field_key": "internal_length_mm", "context": "proposed", "value": "400",
                "unit": "mm", "source_classification": "uploaded_fact",
                "evidence_reference": "DRAWING-02", "validation_status": "valid",
            }],
            "specification_tolerances": [{
                **self._valid_tolerance(), "nominal": "400", "minimum": "398", "maximum": "402"
            }],
        }
        project = {
            "project_id": "P-1", "project_name": "Case", "category": "corrugated",
            "annual_volume": 1000, "currency": "INR",
        }
        normalized = normalize_user_dataset(raw, project)
        value = normalized["intake_values"][0]
        tolerance = normalized["specification_tolerances"][0]
        self.assertEqual(value["value"], 400)
        self.assertEqual(value["source_classification"], "uploaded_fact")
        self.assertEqual(value["evidence_reference"], "DRAWING-02")
        self.assertEqual(tolerance["nominal"], 400)
        self.assertEqual(tolerance["minimum"], 398)
        self.assertEqual(tolerance["maximum"], 402)

    def test_excel_template_inherits_new_baseline_and_proposed_fields(self):
        from io import BytesIO
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(generate_workbook("corrugated", "Cost reduction", "Size optimization")))
        for sheet_name in ("BASELINE", "PROPOSED"):
            field_keys = {workbook[sheet_name].cell(row, 1).value for row in range(2, workbook[sheet_name].max_row + 1)}
            self.assertIn("box_style", field_keys)
            self.assertIn("internal_length_mm", field_keys)
            self.assertIn("regulatory_markings", field_keys)

    def test_model_does_not_generate_approval_or_compression_formula(self):
        fields = {field.key for field in default_registry().get("corrugated").fields}
        self.assertNotIn("approval_decision", fields)
        self.assertNotIn("predicted_bct_n", fields)


if __name__ == "__main__":
    unittest.main()
